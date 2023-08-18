import json
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class TaskTypes:
    SELECT = 'select'
    FREETEXT = 'freeText'
    APPEAL = 'appeal'
    SPEECH = 'speech'


def parseSheetAndSaveAsJson(cell_range_start, cell_range_stop):
    wb: Workbook = load_workbook('files/data.xlsx')
    ws: Worksheet = wb[wb.sheetnames[0]]
    cell_range = ws[cell_range_start:cell_range_stop]

    response = []

    for i in cell_range:
        value_0 = i[0].value
        if value_0 == '' or value_0 is None:
            continue

        value_2 = i[2].value

        page_id = str(value_0) if type(value_0) is not float else str(int(value_0))
        task_type = i[1].value
        next_page = value_2 if type(value_2) is not float else str(int(value_2))
        answer_list = []
        question = ''

        if task_type == TaskTypes.SELECT:
            # if answer list is none so just buttons go next and go prev
            temp_answer_list = i[4].value
            if temp_answer_list is not None:
                next_page = ''

                temp = temp_answer_list.split(';')
                answer_list = [{'text': i.split('=')[0], 'go_to': i.split('=')[1]} for i in temp if ';' in i]


        elif task_type == TaskTypes.APPEAL:
            if i[5].value is not None:
                question += i[5].value

            pass

        response.append({
            "id": str(page_id),
            "type_task": task_type,
            "next_id": str(next_page) if next_page is not None else '',
            "answer_list": answer_list,
            "question": question if question is not None else '',
        })

    with open('files/data.json', 'w', encoding='utf-8') as f:
        json.dump(response, f, ensure_ascii=True, sort_keys=True, indent=4)
